from django.shortcuts import render
from datetime import datetime
from pytz import UTC
from openpyxl import load_workbook
import pandas as pd


# Create your views here.


def index(request, code):
    if request.POST:

        data = {
            "code": code,
            "user-agent": request.POST["user-agent"],
            "ip": request.POST["ip"],
            "city": request.POST["city"],
            "region": request.POST["region"],
            "country": request.POST["country"],
            "timezone": request.POST["timezone"],
            "os": request.POST["os"],
            "time_stamp": str(datetime.now(tz=UTC)),
        }
        print(data)
        add_scanned_records(data)
        return render(request, "thankyou.html", {})

    return render(request, "index.html", {})


def add_scanned_records(data):
    book = None
    try:
        book = load_workbook('./static/TestScan.xlsx')
    except FileNotFoundError:
        pass

    if book:
        # Get the active worksheet
        sheet_obj = book.active
        # Get the max row
        max_row = sheet_obj.max_row
        # Update the cell value
        sheet_obj.cell(row=max_row + 1, column=1, value=max_row-1,)
        # Update values for each cell
        for i in range(2, 11):
            k = list(data.keys())[i-2]
            sheet_obj.cell(row=max_row+1, column=i, value=data[k])

        book.save('./static/TestScan.xlsx')

    else:
        # Create new book
        writer = pd.ExcelWriter('./static/TestScan.xlsx', engine='openpyxl')
        data = pd.DataFrame(data=[data])
        data.to_excel(writer, "Main")
        writer.save()





